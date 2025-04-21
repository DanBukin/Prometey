"""
Модуль с основными вычислительными функциями для моделирования процессов регенеративного охлаждения ЖРД.

Здесь реализованы:
- Расчёт температур охладителя вдоль канала (`find_temp` и его вариации),
- Расчёт температур стенок (`find_temp_st_g`),
- Расчёт тепловых потоков (`find_q_kon`, `find_q_l`, `find_q_sum`),
- Расчёт параметров потока газа (λ, β, и т.д.),
- Интерполяция термодинамических свойств охладителей,
- Обработка координат сопла и других геометрических данных.

Модуль является ядром инженерных расчётов и используется другими модулями проекта.
"""
import pandas as pd
from openpyxl import load_workbook
import math
import numpy as np
import os
from scipy.optimize import root_scalar
from scipy.interpolate import griddata, RBFInterpolator, interp1d
from scipy.spatial import Delaunay
def find_target_row(df, target_string):
    """
        Поиск индекса строки в DataFrame, содержащей заданную строку.

        Args:
            df (pd.DataFrame): Excel-таблица, считанная через pandas.
            target_string (str): Строка, которую нужно найти в таблице.

        Returns:
            int or None: Индекс найденной строки, либо None, если не найдена.
    """
    for index, row in df.iterrows():
        if any(str(cell).strip() == target_string for cell in row):
            return index
    return None
def determine_phi_pr(eta_pr):
    """
        Определяет значение phi_pr по значению eta_pr согласно заданным правилам.

        Args:
            eta_pr (float): Значение эффективности eta_pr.

        Returns:
            float: Соответствующее значение phi_pr.
    """

    if eta_pr <= 5:
        return 0.925
    elif eta_pr <= 15:
        return 0.75
    elif eta_pr <= 20:
        return 0.65
    return 1
def extract_k_l_coefficients(sheet, result):
    """
        Извлекает коэффициенты K2–K6 и L2–L5 из Excel-листа и добавляет их в словарь результата.

        Args:
            sheet (openpyxl.worksheet.worksheet.Worksheet): Активный лист Excel.
            result (dict): Словарь, в который будут добавлены коэффициенты.
    """
    for row_idx, key in zip([2, 3, 4, 5], ["K2", "K3", "K4", "K5"]):
        val = sheet.cell(row=row_idx, column=11).value
        if val is not None and str(val).strip().lower() != "nan":
            result[key] = val
    for row_idx in [6]:
        val = sheet.cell(row=row_idx, column=11).value
        if val is not None and str(val).strip().lower() != "nan":
            result["K6"] = val
    for row_idx, key in zip([2, 3, 4, 5], ["L2", "L3", "L4", "L5"]):
        val = sheet.cell(row=row_idx, column=12).value
        if val is not None and str(val).strip().lower() != "nan":
            result[key] = val
def find_dedal_ikar(file_path_1=None, file_path_2=None):
    """
        Извлекает основные параметры из Excel-файлов, полученных из расчётных комплексов Дедал и Икар.

        Args:
            file_path_1 (str, optional): Путь к файлу Excel из Дедала.
            file_path_2 (str, optional): Путь к файлу Excel из Икара.

        Returns:
            dict: Словарь с ключевыми параметрами (давление, температура, составы, коэффициенты и т.д.).
    """
    result = {}

    if file_path_1:
        df1 = pd.read_excel(file_path_1, header=None)
        idx = find_target_row(df1, 'Основные параметры')
        if idx is None:
            raise ValueError("Не найдены 'Основные параметры' в файле 1")
        offsets = {'p': 4, 'T': 5, 'R': 11, 'k': 14}
        wb1 = load_workbook(file_path_1)
        sheet1 = wb1['Sheet1']
        result.update({key: sheet1.cell(row=idx + offset, column=2).value for key, offset in offsets.items()})
        result["k_m0"] = sheet1.cell(row=6, column=2).value
        result["form_gor"] = sheet1.cell(row=1, column=2).value

        form_gor_map = {
            "{'C': 2, 'H': 8, 'N': 2}": 'НДМГ',
            "{'H': 2}": 'Водород'
        }
        result["cooler_mb"] = form_gor_map.get(result["form_gor"], 'Метан')
        result["H_gor"] = sheet1.cell(row=2, column=2).value
        result["form_ox"] = sheet1.cell(row=3, column=2).value
        result["H_ox"] = sheet1.cell(row=4, column=2).value

    if file_path_2:
        df2 = pd.read_excel(file_path_2, header=None)
        idx_1 = find_target_row(df2, 'Расход ф-ки ок. в пристенке, кг/с')
        idx_2 = find_target_row(df2, 'Соотношение компонентов в пристенке')
        if idx_1 is None:
            raise ValueError("Не найден расход ф-ки ок. в пристенке, кг/с в Икаре")
        if idx_2 is None:
            raise ValueError("Не найдено соотношение компонентов в пристенке в Икаре")

        wb2 = load_workbook(file_path_2)
        sheet2 = wb2['Sheet1']
        result["k_m_pr"] = sheet2.cell(row=idx_1 + 5, column=12).value
        result["eta_pr"] = sheet2.cell(row=idx_2, column=3).value
        result["m_gor"] = round(
            sum(float(sheet2.cell(row=idx_2 + delta, column=6).value) for delta in [-4, -3, 0, 1]),
            3
        )
        result["phi_pr"] = determine_phi_pr(result["eta_pr"])
        extract_k_l_coefficients(sheet2, result)

    return result
def find_index_x_y(file_path):
    """
        Определяет диапазон строк в Excel-файле, содержащих числовые координаты.
        Поиск начинается от строки, содержащей 'Погрешность, %', и продолжается вниз,
        пока строки содержат числовые данные.

        Args:
            file_path (str): Путь к Excel-файлу.

        Returns:
            tuple: (start_row_index, end_row_index) — индексы начала и конца координат.
                   Если "Погрешность, %" не найдена, возвращается (None, None).
    """
    df = pd.read_excel(file_path, header=None)

    # Поиск строки с "Погрешность, %"
    target_row_index = next(
        (idx for idx, row in df.iterrows()
         if any(str(cell).strip() == 'Погрешность, %' for cell in row)),
        None
    )

    if target_row_index is None:
        return None, None

    start_row_index = target_row_index + 4

    # Поиск конца диапазона координат
    end_row_index = len(df)
    for idx in range(start_row_index, len(df)):
        row = df.iloc[idx]
        if not pd.to_numeric(row, errors='coerce').notna().any():
            end_row_index = idx - 1
            break

    return start_row_index, end_row_index
def save_coord_x_y(start_row_index, end_row_index,file_path):
    """
        Извлекает координаты X и Y из Excel-файла и добавляет промежуточные точки
        между первыми двумя отрезками с равномерным шагом.

        Алгоритм:
        1. Первая точка (X=0, Y=Y0) нормализуется относительно первой X-координаты.
        2. Между первой и второй точками добавляются 50 равномерных точек.
        3. Вторая точка (X=50 мм) добавляется явно.
        4. Между второй и третьей точками (из Excel) вставляется 100 точек.
        5. Третья точка из Excel добавляется.
        6. Далее идут остальные координаты из Excel.

        Args:
            start_row_index (int): Индекс строки, с которой начинаются координаты в Excel.
            end_row_index (int): Индекс строки, на которой заканчиваются координаты в Excel.
            file_path (str): Путь к Excel-файлу.

        Returns:
            tuple: Списки координат X и Y в метрах.
    """
    wb = load_workbook(file_path)
    sheet = wb['Sheet1']

    X, Y = [], []

    x0 = sheet.cell(row=start_row_index, column=1).value / 1000  # мм → м
    y0 = sheet.cell(row=start_row_index, column=2).value / 1000

    X.append(0.0)
    Y.append(y0)

    x2 = 0.050  # 50 мм
    step = x2 / 51
    X += [i * step for i in range(1, 51)]
    Y += [y0] * 50
    X.append(x2)
    Y.append(y0)

    x3 = sheet.cell(row=start_row_index + 1, column=1).value / 1000 - x0
    y3 = sheet.cell(row=start_row_index + 1, column=2).value / 1000
    step = (x3 - x2) / 101
    X += [x2 + i * step for i in range(1, 101)]
    Y += [y0] * 100
    X.append(x3)
    Y.append(y3)

    for i in range(start_row_index + 2, end_row_index + 1):
        xi = sheet.cell(row=i, column=1).value / 1000 - x0
        yi = sheet.cell(row=i, column=2).value / 1000
        X.append(xi)
        Y.append(yi)

    return X, Y
def filtred_array(arr,arr_2):
    """
        Фильтрует массив arr, удаляя подряд идущие одинаковые элементы.
        Одновременно удаляются соответствующие элементы из arr_2.

        Args:
            arr (list): Список значений, в котором нужно удалить подряд дублирующиеся элементы.
            arr_2 (list): Сопутствующий список значений, синхронный с arr.

        Returns:
            tuple: (filtered_arr, filtered_arr_2) — списки без повторов по arr.
    """
    result = []
    result_2 = []
    for i in range(len(arr)):
        if i == 0 or arr[i] != arr[i - 1]:
            result.append(arr[i])
            result_2.append(arr_2[i])
    return result, result_2
def okr_2(x):
    """
    Округляет число вниз до ближайшего чётного целого числа.

    Args:
        x (float): Входное число.

    Returns:
        int: Ближайшее чётное число, не больше x.
    """
    x_floor = math.floor(x)
    return x_floor if x_floor % 2 == 0 else x_floor - 1
def find_n_min(beta_reb, t_n_min, d_kp, delta_st, h):
    """
    Вычисляет минимальное количество расчётных сечений вдоль канала охлаждения (n_min),
    округлённое вниз до ближайшего чётного числа.

    Формула:
        D₀.₅ = d_kp * (1 + (2 * delta_st + h) / d_kp)
        n = π * D₀.₅ * cos(beta_reb) / (t_n_min / 1000)
        n_min = округление вниз до ближайшего чётного

    Args:
        beta_reb (float): Угол ребра в радианах.
        t_n_min (float): Минимальный шаг сечения (в миллиметрах).
        d_kp (float): Внутренний диаметр канала охлаждения (в метрах).
        delta_st (float): Толщина стенки (в метрах).
        h (float): Высота канала (в метрах).

    Returns:
        int: Минимальное чётное количество расчётных сечений (n_min).
    """
    D_05_kp = d_kp * (1 + (2 * delta_st + h) / d_kp)
    n_r_kp_min_0 = (math.pi * D_05_kp * math.cos(beta_reb)) / (t_n_min / 1000)
    n_r_kp_min = okr_2(n_r_kp_min_0)
    return n_r_kp_min
def find_n_r_array(X, Y, n_r_kp_min, beta_reb, delta_st, h, t_n_max):
    """
    Вычисляет массив количества расчётных сечений `n_r` по координате Y,
    в зависимости от диаметра канала (пропорционального 2*r) и правила дискретизации.

    Пороговые значения D рассчитываются по формуле:
        D_i = ((t_n_max * multiplier * n_r_kp_min) / (π * cos(beta))) - (2 * delta_st + h)

    Количество сечений увеличивается ступенчато по мере роста диаметра.

    Args:
        X (list of float): Координаты X (не используются в расчёте, но передаются по интерфейсу).
        Y (list of float): Координаты Y (в метрах), по которым будет определяться диаметр.
        n_r_kp_min (int): Минимальное количество сечений.
        beta_reb (float): Угол ребра в радианах.
        delta_st (float): Толщина стенки (в метрах).
        h (float): Высота канала (в метрах).
        t_n_max (float): Максимальный шаг сечения (в миллиметрах).

    Returns:
        list of int: Массив количества сечений n_r для каждой точки по Y.
    """
    multipliers = [1, 2, 4, 8, 16]
    D = [
        (t_n_max * m * n_r_kp_min / 1000) / (math.pi * math.cos(beta_reb)) - (2 * delta_st + h)
        for m in multipliers
    ]

    n_r_array = []
    for r in Y:
        diameter = 2 * r
        if diameter < D[0]:
            n_r_array.append(n_r_kp_min)
        elif diameter < D[1]:
            n_r_array.append(2 * n_r_kp_min)
        elif diameter < D[2]:
            n_r_array.append(4 * n_r_kp_min)
        elif diameter < D[3]:
            n_r_array.append(8 * n_r_kp_min)
        elif diameter < D[4]:
            n_r_array.append(16 * n_r_kp_min)
        else:
            n_r_array.append(32 * n_r_kp_min)

    return n_r_array
def equation(lambda_, k, F_otn_i):
    """
    Расчёт значения функции для подбора параметра расширения потока λ (лямбда) по заданному отношению площадей.

    Эта функция определяет выражение, корень которого соответствует искомому значению λ в уравнении
    для изэнтропического истечения газа через сопло:

        λ · [1 - λ² · ((k - 1)/(k + 1))]^(1 / (k - 1)) · [(k + 1)/2]^(1 / (k - 1)) = 1 / F_otn_i

    Где:
        - λ — отношение локальной скорости потока к критической (расширение),
        - k — показатель адиабаты газа (обычно 1.2–1.4),
        - F_otn_i — относительное отношение площадей F/F_кр в рассматриваемом сечении сопла.

    Возвращаемое значение — левая часть уравнения минус правая,
    используется при численном решении (например, методом Брента или Ньютона) для нахождения корня.

    Особенности:
    - Возвращает `np.nan`, если выражение под корнем отрицательно или данные некорректны.
    - Предусмотрена защита от деления на ноль и комплексных чисел (возведение в степень дробной степени).

    Args:
        lambda_ (float): Предполагаемое значение коэффициента расширения потока.
        k (float): Показатель адиабаты рабочего тела.
        F_otn_i (float): Относительное отношение площадей F / F_кр для данного сечения.

    Returns:
        float: Значение функции (левая часть минус правая), либо np.nan при ошибке или недопустимом λ.
    """
    try:
        if lambda_ <= 0 or F_otn_i <= 0 or k <= 1:
            return np.nan

        term2_base = 1 - (lambda_ ** 2) * ((k - 1) / (k + 1))
        if term2_base <= 0:
            return np.nan  # избегаем комплексных значений

        term2 = term2_base ** (1 / (k - 1))
        term3 = ((k + 1) / 2) ** (1 / (k - 1))

        return float(lambda_ * term2 * term3 - (1 / F_otn_i))

    except Exception as e:
        print(f"Ошибка в equation(): {e}, λ={lambda_}, k={k}, F={F_otn_i}")
        return np.nan
def solve_lambda(F_i, k, lower=0.01, upper=4.0, xtol=1e-6):
    """
    Надёжный численный подбор коэффициента расширения потока λ методом Брента.

    Функция решает изэнтропическое уравнение истечения газа из сопла (см. функцию `equation(...)`) для заданного
    отношения площадей F/F_кр и показателя адиабаты k, возвращая соответствующее значение λ — безразмерного параметра,
    отражающего степень расширения потока в данном сечении сопла.

    Реализация использует метод `brentq` из `scipy.optimize.root_scalar`, при этом:
    - Проверяется допустимость интервала поиска (`bracket`) — знак функции на концах должен быть разным;
    - Выполняется автоматическое ограничение `upper`, чтобы избежать комплексных значений в уравнении (по λ² < (k + 1)/(k - 1));
    - Если корень находится точно на границе — возвращается `lower` или `upper`;
    - При возникновении ошибок возвращается `None`.

    Args:
        F_i (float): Отношение площадей F / F_кр в конкретной точке сопла.
        k (float): Показатель адиабаты газа.
        lower (float): Левая граница поиска λ (по умолчанию 0.01).
        upper (float): Правая граница поиска λ (по умолчанию 4.0).
        xtol (float): Абсолютная точность подбора корня (по умолчанию 1e-6).

    Returns:
        float or None: Найденное значение λ, соответствующее заданному F_i и k, либо None при сбое или недопустимом интервале.
    """
    try:
        lambda_max = math.sqrt((k + 1) / (k - 1))
        upper = min(upper, lambda_max - 1e-6)

        f_a = equation(lower, k, F_i)
        f_b = equation(upper, k, F_i)

        # ✅ Обработка корня на границе
        if abs(f_a) < 1e-8:
            return lower
        if abs(f_b) < 1e-8:
            return upper

        if any([
            isinstance(f_a, complex), isinstance(f_b, complex),
            np.isnan(f_a), np.isnan(f_b),
            f_a * f_b > 0
        ]):
            print(f"⚠️ Неправильный интервал: F={F_i}, f(a)={f_a}, f(b)={f_b}")
            return None

        sol = root_scalar(
            equation,
            args=(k, F_i),
            bracket=[lower, upper],
            method='brentq',
            xtol=xtol
        )

        if not sol.converged:
            print(f"⚠️ Корень не найден точно для F={F_i}")
        return sol.root

    except Exception as e:
        print(f"❌ Ошибка в solve_lambda: {e}")
        return None
def find_q_kon(number, Y, c_p_T_0g, c_p_T_st, T_st_g, T_og, mu_og, R_og, D_otn, d_kp, F_otn, b, p_k, k, iter):
    """
    Расчёт теплового потока q к стенке канала ЖРД с учётом параметров газа, геометрии и температурных условий.

    Args:
        number (list of int): Индексы расчётных точек.
        Y (list of float): Радиальные координаты точек.
        c_p_T_0g (float): Теплоёмкость газа при температуре газа.
        c_p_T_st (list of float): Теплоёмкость в стенке (по точкам).
        T_st_g (list of float): Температура стенки газа (по точкам).
        T_og (float): Температура газа.
        mu_og (float): Динамическая вязкость газа.
        R_og (float): Газовая постоянная для газа.
        D_otn (list of float): Относительный диаметр в точках.
        d_kp (float): Диаметр канала.
        F_otn (list of float): Относительное сечение.
        b (float): Ширина щели охлаждения (в мм).
        p_k (float): Давление в камере (в МПа).
        k (float): Показатель адиабаты.
        iter (int): Итерируемая переменная (не используется в логике, возможно зарезервирована).

    Returns:
        tuple: (q, betta, T_st_otn, lambd, S_1)
            q (list of float): Тепловой поток к стенке.
            betta (list of float): Коэффициент скорости.
            T_st_otn (list of float): Относительная температура стенки.
            lambd (list of float): Коэффициент расширения потока.
            S_1 (list of float): Вспомогательная характеристика теплоотдачи.
    """
    lambd = []
    betta = []

    for i in number:
        F_i = F_otn[i]

        if i <= Y.index(min(Y)):  # До критики → дозвук
            lambd_i = solve_lambda(F_i, k, lower=0.01, upper=1.0)
        else:  # После критики → сверхзвук
            lambd_i = solve_lambda(F_i, k, lower=1.0, upper=4.0)

        if lambd_i is None:
            lambd_i = 0  # или np.nan, или continue
        lambd.append(lambd_i)

        betta.append(lambd_i * math.sqrt((k - 1) / (k + 1)))

    c_p_sr = [0.5 * (c_p_T_0g + c_p_st_i) for c_p_st_i in c_p_T_st]

    T_st_otn = [T_st / T_og for T_st in T_st_g]

    S_1 = [
        (2.065 * c_p_sr[i] * (T_og - T_st_g[i]) * (mu_og ** 0.15)) /
        (((R_og * T_og) ** 0.425) *
         ((1 + T_st_otn[i]) ** 0.595) *
         ((3 + T_st_otn[i]) ** 0.15))
        for i in range(len(number))
    ]

    slab_alpha = 1.813 * ((2 / (k + 1)) ** (0.85 / (k - 1))) * ((2 * k / (k + 1)) ** 0.425)

    slab_zed = [
        1.769 * ((1 - betta[i] ** 2 + betta[i] ** 2 *
                  (1 - 0.086 * (1 - betta[i] ** 2) /
                   (1 - T_st_otn[i] - 0.1 * betta[i] ** 2))) /
                 (1 - T_st_otn[i] - 0.1 * betta[i] ** 2)) ** 0.54
        for i in range(len(number))
    ]

    A = 0.01352
    B = [0.4842 * A * slab_alpha * slab_zed[i] for i in range(len(number))]

    q = [
        b * 1e-6 * B[i] * (((1 - betta[i] ** 2) * (p_k * 1e6) ** 0.85 * S_1[i]) /
                          (D_otn[i] ** 1.82 * d_kp ** 0.15 * 0.75 ** 0.58))
        for i in range(len(number))
    ]

    return q, betta, T_st_otn, lambd, S_1
def find_l_e_D(M_co2, M, p_k, M_h2o, X, Y, R_k, T_k):
    """
    Вычисляет степень чёрнотелости продуктов сгорания (eps_g) на основе
    геометрии канала и состава газов (CO2 и H2O).

    Args:
        M_co2 (float): Массовая доля CO2.
        M (float): Общая масса продуктов сгорания.
        p_k (float): Давление в камере (в МПа).
        M_h2o (float): Массовая доля H2O.
        X (list of float): Список координат X.
        Y (list of float): Список координат Y.
        R_k (float): Газовая постоянная.
        T_k (float): Температура стенки (в K).

    Returns:
        float: Эффективная степень чёрнотелости газов (eps_g).
    """
    p_co2 = (M_co2 / M) * p_k
    p_h2o = (M_h2o / M) * p_k

    x_value = X[1] / (2 * Y[0])
    x = np.array([1, 1.5, 2.5, 4])
    y = np.array([0.6, 0.75, 0.85, 0.9])
    y_value = np.interp(x_value, x, y, left=y[0], right=y[-1])
    L_e = y_value * 2 * Y[0]

    ro_h2o = (p_h2o * 1e6) / (R_k * T_k)
    ro_h2o_l_e = ro_h2o * L_e
    lg_ro_h2o_l_e = math.log10(ro_h2o_l_e)

    eps_0_h2o = 0.14
    beta_h2o = 1.86
    eps_h2o = beta_h2o * eps_0_h2o

    return eps_h2o
def find_q_l(eps_g, T_k, phi, D_otn, number, Y):
    """
    Расчёт лучистого теплового потока (q_l) в разных сечениях канала охлаждения.

    Args:
        eps_g (float): Эффективная степень чёрнотелости продуктов сгорания.
        T_k (float): Температура стенки (K).
        phi (float): Коэффициент относительной площади.
        D_otn (list of float): Список относительных диаметров.
        number (list of int): Индексы расчётных сечений.
        Y (list of float): Радиальные координаты.

    Returns:
        list of float: Список лучистых тепловых потоков [МВт/м²].
    """
    eps_st = 0.8
    eps_st_ef = 0.5 * (eps_st + 1)
    c_0 = 5.67  # Постоянная Стефана-Больцмана * 10⁻⁸

    q_l_km = eps_st_ef * eps_g * c_0 * ((T_k / 100) ** 4)
    q_l_kc = q_l_km * phi

    comparison_number = 1.2
    index_of_element = next((idx for idx, val in enumerate(D_otn) if val < comparison_number), -1)

    q_l = []
    for i in number:
        if i <= 51:
            # Линейный рост между 0.25 и полной мощностью на 51 точках
            base = q_l_km * 0.25 * 1e-6
            delta = (q_l_km * phi * 1e-6) - base
            q_l.append(base + delta * (i / 51))
        elif i < index_of_element:
            q_l.append(q_l_km * phi * 1e-6)
        elif i >= index_of_element and i < Y.index(min(Y)):
            q_l.append(q_l_kc * 1e-6 * (1 - 12.5 * ((1.2 - D_otn[i]) ** 2)))
        elif i == Y.index(min(Y)):
            q_l.append(0.5 * q_l_kc * 1e-6)
        else:
            q_l.append((0.5 * q_l_kc * 1e-6) / (D_otn[i] ** 2))
    return q_l
def find_q_sum(q_kon, q_l):
    """
    Суммирует конвективный и лучистый тепловые потоки по каждому сечению.

    Args:
        q_kon (list of float): Конвективные тепловые потоки.
        q_l (list of float): Лучистые тепловые потоки.

    Returns:
        list of float: Суммарные тепловые потоки (конвекция + излучение).
    """
    return [k + l for k, l in zip(q_kon, q_l)]
def find_temp_slozh_2(ind_peret,delta_S,q_sum,a,number,p_ohl,m_ohl,T_nach,name_ohl,T_aray,p_aray,C_aray):
    """
    Вычисляет распределение температуры охладителя при сложной системе охлаждения:
    поток сначала движется к середине канала, затем поворачивает и идёт обратно (противотоком).

    Модель делит канал на две части:
    - Прямая часть (от камеры к середине) рассчитывается напрямую.
    - Обратная часть (от середины к выходу сопла) рассчитывается с начальной температурой,
      равной финальной температуре первой части.

    Args:
        ind_peret (int): Индекс точки разворота потока (середина канала).
        delta_S (list of float): Площадь сечения канала по длине (м²).
        q_sum (list of float): Суммарный тепловой поток на участке [Вт/м²].
        a (float): Коэффициент теплообмена.
        number (list of int): Индексы расчётных точек.
        p_ohl (list of float): Давление охладителя по участкам (Па).
        m_ohl (float): Массовый расход охладителя (кг/с).
        T_nach (float): Начальная температура охладителя (К).
        name_ohl (str): Название охладителя (например, "водород", "метан").
        T_aray (list of float): Температурный массив для интерполяции.
        p_aray (list of float): Давления для интерполяции теплоёмкости.
        C_aray (list of float): Таблица теплоёмкости (по T и p).

    Returns:
        tuple:
            - T_final (np.ndarray): Распределение температуры по всей длине канала.
            - C_p_ohl_final (np.ndarray): Теплоёмкость охладителя в каждой точке.
            - C_p_raznitsa_final (np.ndarray): Изменение теплоёмкости между точками.

    Примечание:
        Использует вспомогательную функцию `find_temp()` для интегрирования температуры вдоль канала.
    """
    mid_index = ind_peret
    # --- 🔹 1. По потоку (от камеры к середине) ---
    number_part1=number[:mid_index]
    delta_S_part1 = delta_S[:mid_index]  # Оставляем вторую половину
    p_ohl_part1 = p_ohl[:mid_index]
    q_sum_part1 = np.array(q_sum[:mid_index])  # Оставляем вторую половину
    T_part1, C_p_ohl_part1, C_p_raznitsa_part1 = find_temp(
        number_part1, delta_S_part1, p_ohl_part1, q_sum_part1, m_ohl, a, T_nach,name_ohl,T_aray,p_aray,C_aray, reverse=False)
    T_middle = T_part1[-1]

    # --- 🔹 2. Противоток (от середины к выходу сопла) ---
    number_part2 = number[mid_index:]
    delta_S_part2 = delta_S[mid_index:]  # Оставляем вторую половину
    q_sum_part2 = np.array(q_sum[mid_index:])  # Оставляем вторую половину
    p_ohl_part2=p_ohl[mid_index:]
    T_part2, C_p_ohl_part2, C_p_raznitsa_part2 = find_temp(
        number_part2, delta_S_part2, p_ohl_part2, q_sum_part2, m_ohl, a, T_middle,name_ohl,T_aray,p_aray,C_aray, reverse=True)
    T_final = np.concatenate((T_part1, T_part2))
    C_p_ohl_final = np.concatenate((C_p_ohl_part1, C_p_ohl_part2))
    C_p_raznitsa_final = np.concatenate((C_p_raznitsa_part1, C_p_raznitsa_part2))
    print(f'Максимальная температура охладителя: {max(T_final)} К, подогрев составляет {max(T_final) - min(T_final):.2f}К')
    return T_final, C_p_ohl_final, C_p_raznitsa_final
def find_temp_slozh_1(ind_peret,delta_S,q_sum,a,number,p_ohl,m_ohl,T_nach,name_ohl,T_aray,p_aray,C_aray):
    """
    Вычисляет распределение температуры охладителя в конфигурации с односторонним охлаждением:
    сначала поток идёт от середины канала к выходу сопла, затем — в обратную сторону к камере.

    Используется модель двух участков:
    - Прямой участок (от середины к соплу) — с начальной температурой `T_nach`.
    - Обратный участок (от середины к началу канала) — с начальной температурой,
      равной финальной температуре прямого участка.

    Args:
        ind_peret (int): Индекс точки разворота (середина канала).
        delta_S (list of float): Площадь сечения канала по длине (м²).
        q_sum (list of float): Суммарный тепловой поток на каждом участке (Вт/м²).
        a (float): Коэффициент теплоотдачи (Вт/м²·К).
        number (list of int): Индексы расчётных точек.
        p_ohl (list of float): Давление охладителя (Па) на каждом участке.
        m_ohl (float): Массовый расход охладителя (кг/с).
        T_nach (float): Начальная температура охладителя (К).
        name_ohl (str): Название охладителя (например, "водород", "метан").
        T_aray (list of float): Температуры для интерполяции теплоёмкости.
        p_aray (list of float): Давления для интерполяции теплоёмкости.
        C_aray (list of float): Таблица теплоёмкости (по T и p).

    Returns:
        tuple:
            - T_final (np.ndarray): Температура охладителя вдоль всей длины канала.
            - C_p_ohl_final (np.ndarray): Теплоёмкость охладителя в каждой точке.
            - C_p_raznitsa_final (np.ndarray): Изменение теплоёмкости между точками.

    Примечание:
        Основана на вызове функции `find_temp`, которая производит численное интегрирование температуры.
    """
    mid_index = ind_peret
    # --- 🔹 2. По потоку (от середины к выходу сопла) ---
    delta_S_fwd = delta_S[mid_index:]  # Оставляем первую половину
    q_sum_fwd = np.array(q_sum[mid_index:])  # Масштабируем

    T_part2, C_p_ohl_part2, C_p_raznitsa_part2 = find_temp(
        number[mid_index:], delta_S_fwd, p_ohl[mid_index:], q_sum_fwd, m_ohl, a, T_nach,name_ohl,T_aray,p_aray,C_aray, reverse=False)
    T_middle = T_part2[-1]
    # --- 🔹 1. Противоток (от середины к началу сопла) ---
    T_part1, C_p_ohl_part1, C_p_raznitsa_part1 = find_temp(
        number[:mid_index], delta_S[:mid_index], p_ohl[:mid_index], q_sum[:mid_index], m_ohl, a, T_middle,name_ohl,T_aray,p_aray,C_aray, reverse=True)
    # Температура в середине сопла после первой стадии
    # Объединяем два участка в один массив
    T_final = np.concatenate((T_part1, T_part2))
    C_p_ohl_final = np.concatenate((C_p_ohl_part1, C_p_ohl_part2))
    C_p_raznitsa_final = np.concatenate((C_p_raznitsa_part1, C_p_raznitsa_part2))

    print(f'Максимальная температура охладителя: {max(T_final)} К, подогрев составляет {max(T_final) - min(T_final):.2f}К')
    return T_final, C_p_ohl_final, C_p_raznitsa_final
def find_temp_slozh_5(m_ohl_1,m_ohl_2,index_peret_1,index_peret_2,delta_S,q_sum,number,p_ohl,a,T_nach,name_ohl,T_aray,p_aray,C_aray):
    """
    Вычисляет распределение температуры охладителя при сложной схеме охлаждения с двумя потоками:
    один — прямоточный (от середины к выходу сопла), второй — противоточный (от сопла к камере).
    Потоки имеют разные массовые расходы.

    Структура:
    🔹 Часть 1: Противоток от `index_peret_2` к началу канала — с массовым расходом `m_ohl_2`.
    🔹 Часть 2: Прямоток от `index_peret_2` к выходу сопла — с массовым расходом `m_ohl_1`.

    Args:
        m_ohl_1 (float): Массовый расход в прямоточном участке (к соплу), кг/с.
        m_ohl_2 (float): Массовый расход в противоточном участке (к камере), кг/с.
        index_peret_1 (int): Индекс разворота прямотока (не используется в этой реализации).
        index_peret_2 (int): Индекс разворота противотока и начала прямотока.
        delta_S (list of float): Площади сечения канала вдоль длины (м²).
        q_sum (list of float): Суммарный тепловой поток на участке (Вт/м²).
        number (list of int): Индексы расчётных точек.
        p_ohl (list of float): Давление охладителя (Па).
        a (float): Коэффициент теплоотдачи (Вт/м²·К).
        T_nach (float): Начальная температура охладителя (К).
        name_ohl (str): Название охладителя.
        T_aray (list of float): Температуры для интерполяции теплоёмкости.
        p_aray (list of float): Давления для интерполяции теплоёмкости.
        C_aray (list of float): Таблица теплоёмкости по T и p.

    Returns:
        tuple:
            - T_final (np.ndarray): Температурный профиль охладителя.
            - C_p_ohl_final (np.ndarray): Теплоёмкость охладителя в каждой точке.
            - C_p_raznitsa_final (np.ndarray): Изменение теплоёмкости между точками.
            - m_final (np.ndarray): Массив массовых расходов вдоль канала.
    """
    T_nach_0 = T_nach
    # --- 🔹 1. Противоток (от index_peret_2 к камере) ---
    T_part2, C_p_ohl_part2, C_p_raznitsa_part2 = find_temp(number[:index_peret_2],
                                                           delta_S[:index_peret_2],
                                                           p_ohl[:index_peret_2],
                                                           q_sum[:index_peret_2], m_ohl_2, a, T_nach,
                                                           name_ohl, T_aray, p_aray, C_aray, reverse=True)
    T_middle_2 = T_part2[0]
    m_ohl_array_2 = [m_ohl_2] * len(T_part2)
    # --- 🔹 2. По потоку (от index_peret_2 к концу зеркальной части) --
    delta_S_fwd = delta_S[index_peret_2:]  # Оставляем первую половину
    m_ohl_array_1 = [m_ohl_1] * len(delta_S_fwd)
    q_sum_fwd = np.array(q_sum[index_peret_2:])  # Масштабируем
    T_part1, C_p_ohl_part1, C_p_raznitsa_part1 = find_temp(number[index_peret_2:], delta_S_fwd, p_ohl[index_peret_2:],
                                                           q_sum_fwd, m_ohl_1, a, T_nach_0, name_ohl, T_aray, p_aray,
                                                           C_aray, reverse=False)

    T_final = np.concatenate((T_part2, T_part1))
    m_final = np.concatenate((m_ohl_array_2, m_ohl_array_1))
    C_p_ohl_final = np.concatenate((C_p_ohl_part2, C_p_ohl_part1))
    C_p_raznitsa_final = np.concatenate((C_p_raznitsa_part2, C_p_raznitsa_part1))
    print(f'Максимальная температура охладителя: {max(T_final)} К, подогрев составляет {max(T_final) - min(T_final):.2f}К')
    return T_final, C_p_ohl_final, C_p_raznitsa_final,m_final
def find_temp_slosh_tema(m_ohl_1,m_ohl_2,index_peret_1,index_peret_2,delta_S,q_sum,number,p_ohl,a,T_nach,name_ohl,T_aray,p_aray,C_aray):
    """
    Модель расчёта температуры охладителя для трёхпоточной схемы:
    ▸ Прямоток от разворота 1 к соплу,
    ▸ Противоток от разворота 2 к развороту 1,
    ▸ Суммарный противоток от камеры до разворота 2, где два потока объединяются.

    Сначала считается температура в двух потоках независимо, затем рассчитывается
    результирующий объединённый поток с учётом энтальпийной равновесной температуры.

    Args:
        m_ohl_1 (float): Массовый расход охладителя в первом (прямоточном) потоке [кг/с].
        m_ohl_2 (float): Массовый расход во втором (противоточном) потоке [кг/с].
        index_peret_1 (int): Индекс точки поворота первого потока (прямоток).
        index_peret_2 (int): Индекс точки поворота второго потока (противоток).
        delta_S (list of float): Площади каналов вдоль длины (м²).
        q_sum (list of float): Суммарный тепловой поток в каждой точке [Вт/м²].
        number (list of int): Индексы расчётных точек.
        p_ohl (list of float): Давление охладителя в каждой точке [Па].
        a (float): Коэффициент теплоотдачи (Вт/м²·К).
        T_nach (float): Начальная температура охладителя (К).
        name_ohl (str): Название охладителя (напр. "метан", "водород").
        T_aray (list of float): Температурный массив для интерполяции.
        p_aray (list of float): Давления для интерполяции теплоёмкости.
        C_aray (list of float): Таблица значений теплоёмкости (по T и p).

    Returns:
        tuple:
            - T_final (np.ndarray): Полный температурный профиль по всей длине канала.
            - C_p_ohl_final (np.ndarray): Теплоёмкость охладителя в каждой точке.
            - C_p_raznitsa_final (np.ndarray): Изменение теплоёмкости между точками.
            - m_final (np.ndarray): Распределение массового расхода вдоль длины канала.
    """
    T_nach_0=T_nach
    # --- 🔹 1. По потоку (от index_peret_1 к выходу сопла) --
    delta_S_fwd = delta_S[index_peret_1:]  # Оставляем первую половину
    m_ohl_array_1=[m_ohl_1]*len(delta_S_fwd)
    q_sum_fwd = np.array(q_sum[index_peret_1:])  # Масштабируем
    T_part1, C_p_ohl_part1, C_p_raznitsa_part1 = find_temp(number[index_peret_1:], delta_S_fwd, p_ohl[index_peret_1:], q_sum_fwd, m_ohl_1, a, T_nach,name_ohl,T_aray,p_aray,C_aray, reverse=False)
    T_middle = T_part1[-1]

    # --- 🔹 2. Противоток (от index_peret_2 к index_peret_1) ---
    T_part2, C_p_ohl_part2, C_p_raznitsa_part2 = find_temp(number[index_peret_2:index_peret_1], delta_S[index_peret_2:index_peret_1], p_ohl[index_peret_2:index_peret_1], q_sum[index_peret_2:index_peret_1], m_ohl_2, a, T_nach,name_ohl,T_aray,p_aray,C_aray, reverse=True)
    T_middle_2 = T_part2[0]
    m_ohl_array_2 = [m_ohl_2] * len(T_part2)
    # --- 🔹 3. Суммарный противопоток (от камеры к index_peret_2) ---
    T_nach_1=((m_ohl_1*T_middle)+(m_ohl_2*T_middle_2))/(m_ohl_1+m_ohl_2)
    T_part3, C_p_ohl_part3, C_p_raznitsa_part3 = find_temp(number[:index_peret_2],
                                                           delta_S[:index_peret_2], p_ohl[:index_peret_2],
                                                           q_sum[:index_peret_2], m_ohl_2+m_ohl_1, a, T_nach_1,name_ohl,T_aray,p_aray,C_aray,
                                                           reverse=True)
    T_middle_3 = T_part3[-1]
    m_ohl_array_3 = [m_ohl_1+m_ohl_2] * len(T_part3)
    print(f'Расчёт противотока завершён!')
    T_final = np.concatenate((T_part3, T_part2,T_part1))
    m_final=np.concatenate((m_ohl_array_3, m_ohl_array_2,m_ohl_array_1))
    C_p_ohl_final = np.concatenate((C_p_ohl_part3, C_p_ohl_part2,C_p_ohl_part1))
    C_p_raznitsa_final = np.concatenate((C_p_raznitsa_part3, C_p_raznitsa_part2,C_p_raznitsa_part1))
    print(f'Максимальная температура охладителя: {max(T_final)} К, подогрев составляет {max(T_final) - min(T_final):.2f}К')
    return T_final, C_p_ohl_final, C_p_raznitsa_final,m_final
def find_temp_slozh_6(m_ohl_1,m_ohl_2,index_peret,delta_S,q_sum,number,p_ohl,a,T_nach_1,T_nach_2,T_aray_1,p_aray_1,C_aray_1,T_aray_2,p_aray_2,C_aray_2):
    """
    Расчёт температуры охладителя при двух независимых противоточных потоках:
    ▸ Первый поток (АТ) идёт от точки разворота `index_peret` к соплу,
    ▸ Второй поток (НДМГ) — от камеры до точки `index_peret`.

    Каждый поток имеет свой массовый расход, начальную температуру и набор термодинамических данных.

    Args:
        m_ohl_1 (float): Массовый расход первого охладителя (АТ), кг/с.
        m_ohl_2 (float): Массовый расход второго охладителя (НДМГ), кг/с.
        index_peret (int): Индекс точки разделения потоков.
        delta_S (list of float): Площади сечения по длине канала (м²).
        q_sum (list of float): Суммарный тепловой поток на участке (Вт/м²).
        number (list of int): Индексы расчётных точек.
        p_ohl (list of float): Давление охладителя (Па) по участкам.
        a (float): Коэффициент теплоотдачи (Вт/м²·К).
        T_nach_1 (float): Начальная температура потока АТ (К).
        T_nach_2 (float): Начальная температура потока НДМГ (К).
        T_aray_1 (list of float): Температуры для интерполяции теплоёмкости АТ.
        p_aray_1 (list of float): Давления для интерполяции теплоёмкости АТ.
        C_aray_1 (list of float): Таблица теплоёмкости АТ (по T и p).
        T_aray_2 (list of float): Температуры для интерполяции теплоёмкости НДМГ.
        p_aray_2 (list of float): Давления для интерполяции теплоёмкости НДМГ.
        C_aray_2 (list of float): Таблица теплоёмкости НДМГ (по T и p).

    Returns:
        tuple:
            - T_final (np.ndarray): Полный температурный профиль по длине канала.
            - C_p_ohl_final (np.ndarray): Теплоёмкость охладителя в каждой точке.
            - C_p_raznitsa_final (np.ndarray): Изменение теплоёмкости между точками.
            - m_final (np.ndarray): Распределение массового расхода вдоль канала.
    """
    # --- 🔹 1. Противоток (от index_peret к соплу) ---
    T_part1, C_p_ohl_part1, C_p_raznitsa_part1 = find_temp(number[index_peret:],
                                                           delta_S[index_peret:],
                                                           p_ohl[index_peret:],
                                                           q_sum[index_peret:], m_ohl_1, a, T_nach_1,
                                                           'АТ', T_aray_1, p_aray_1, C_aray_1, reverse=True)
    m_ohl_array_1 = [m_ohl_1] * len(T_part1)
    # --- 🔹 2. Противоток (от камеры к index_peret) ---
    T_part2, C_p_ohl_part2, C_p_raznitsa_part2 = find_temp(number[:index_peret],
                                                           delta_S[:index_peret],
                                                           p_ohl[:index_peret],
                                                           q_sum[:index_peret], m_ohl_2, a, T_nach_2,
                                                           'НДМГ', T_aray_2, p_aray_2, C_aray_2, reverse=True)
    m_ohl_array_2=[m_ohl_2]*len(T_part2)
    T_final = np.concatenate((T_part2, T_part1))
    m_final = np.concatenate((m_ohl_array_2, m_ohl_array_1))
    C_p_ohl_final = np.concatenate((C_p_ohl_part2, C_p_ohl_part1))
    C_p_raznitsa_final = np.concatenate((C_p_raznitsa_part2, C_p_raznitsa_part1))
    print(f'Максимальная температура охладителя: {max(T_final)} К, подогрев составляет {max(T_final) - min(T_final):.2f}К')
    return T_final, C_p_ohl_final, C_p_raznitsa_final, m_final
def find_temp(number, delta_S, p_ohl, q_sum, m_ohl, a, T_nach,name_ohl, T_aray, p_aray, C_aray, reverse=True, max_iter=100, verbose=False):
    """
    Численно рассчитывает температурный профиль охладителя вдоль канала с учётом теплообмена и переменной теплоёмкости.

    Args:
        number (list of int): Индексы расчётных сечений.
        delta_S (list of float): Площади поперечного сечения (м²).
        p_ohl (list of float): Давление охладителя (Па).
        q_sum (list of float): Суммарный тепловой поток (Вт/м²).
        m_ohl (float): Массовый расход охладителя (кг/с).
        a (float): Коэффициент теплоотдачи (Вт/м²·К).
        T_nach (float): Начальная температура охладителя (К).
        name_ohl (str): Название охладителя (используется для подбора Cp).
        T_aray (list): Температурный массив.
        p_aray (list): Давления.
        C_aray (list): Таблица теплоёмкости Cp(T, p).
        reverse (bool): Направление расчёта (True — от выхода к камере).
        max_iter (int): Максимальное число итераций на одну точку.
        verbose (bool): Печатать отладочную информацию.

    Returns:
        tuple:
            T (list of float): Температурный профиль.
            C_p_ohl (list of float): Cp охладителя.
            C_p_raznitsa (list of float): Ошибка между приближениями.
    """
    if reverse:
        number = number[::-1]
        delta_S = delta_S[::-1][1:]
        q_sum = q_sum[::-1]

    T = [T_nach] + [0] * (len(number) - 1)
    C_p_ohl = [0] * len(number)
    C_p_raznitsa = [0] * len(number)

    q_sum = np.array(q_sum) * a

    for i in range(1, len(number)):
        delta_T_step = 1.0
        for iteration in range(max_iter):
            T_sr_1 = T[i - 1] + 0.5 * delta_T_step
            Cp_1 = interpolate_C(T_sr_1, p_ohl[i - 1], name_ohl, T_aray, p_aray, C_aray)

            delta_T_2 = 0.5 * (q_sum[i - 1] + q_sum[i]) * 1e6 * delta_S[i - 1] / (m_ohl * Cp_1)

            T_sr_2 = T[i - 1] + 0.5 * delta_T_2
            Cp_2 = interpolate_C(T_sr_2, p_ohl[i - 1], name_ohl, T_aray, p_aray, C_aray)

            diff = abs(Cp_2 - Cp_1) * 100 / Cp_1
            if diff <= 5:
                T[i] = T[i - 1] + delta_T_2
                C_p_ohl[i] = Cp_2
                C_p_raznitsa[i] = diff
                break
            delta_T_step += 1

            if iteration == max_iter - 1 and verbose:
                print(f"[WARN] Невозможно достичь сходимости на i={i}")

    # Начальные значения Cp и корректировка направлений
    C_p_ohl[0] = interpolate_C(T_nach, p_ohl[0], name_ohl, T_aray, p_aray, C_aray)
    Cp_1_1 = interpolate_C(T[1], p_ohl[1], name_ohl, T_aray, p_aray, C_aray)
    C_p_raznitsa[0] = abs(Cp_1_1 - C_p_ohl[0]) * 100 / C_p_ohl[0]

    if reverse:
        T = T[::-1]
        C_p_ohl = C_p_ohl[::-1]
        C_p_raznitsa = C_p_raznitsa[::-1]

    return T, C_p_ohl, C_p_raznitsa
def is_inside(points, test_point):
    """
    Проверяет, принадлежит ли тестовая точка выпуклой области, построенной по заданным точкам с помощью триангуляции Делоне.

    Args:
        points (array-like): Массив точек (n x d), где n — количество точек, d — размерность (2D или 3D и т.д.).
        test_point (array-like): Точка, которую нужно проверить (в том же числе измерений, что и points).

    Returns:
        bool: True, если точка лежит внутри выпуклой оболочки множества, иначе False.
    """
    points = np.asarray(points)
    test_point = np.asarray(test_point)
    try:
        tri = Delaunay(points)
        return tri.find_simplex(test_point) >= 0
    except Exception as e:
        raise ValueError(f"Ошибка при проверке принадлежности точки: {e}")
def interpolate_C(T_input, p_input, name_ohl,T_aray, p_aray, C_aray,verbose=False):
    """
    Интерполирует или экстраполирует значение теплоёмкости C по температуре и давлению.

    Для "сложных" компонентов (с зависимостью от давления) используется 2D-интерполяция по (T, p),
    в остальных случаях — линейная 1D-интерполяция по температуре.

    Args:
        T_input (float): Температура, К.
        p_input (float): Давление, Па.
        name_ohl (str): Название охладителя.
        T_aray (array-like): Массив температур.
        p_aray (array-like): Массив давлений (для сложных веществ).
        C_aray (array-like): Таблица значений теплоёмкости.
        verbose (bool): Выводить предупреждения при экстраполяции.

    Returns:
        float: Интерполированное значение C.
    """
    complex_components = {"Водород", "АТ", "Кислород", "Гелий", "Аммиак", "Метан"}
    if name_ohl in complex_components:
        points = np.column_stack((T_aray, p_aray))
        test_point = np.array([T_input, p_input])

        if is_inside(points, test_point):
            C_interp = griddata(points, C_aray, (T_input, p_input), method='cubic')
            if np.isnan(C_interp):
                C_interp = griddata(points, C_aray, (T_input, p_input), method='linear')
        else:
            if verbose:
                print(f"⚠️ Экстраполяция по (T={T_input}, p={p_input}) для '{name_ohl}'")
            rbf = RBFInterpolator(points, C_aray, kernel='thin_plate_spline')
            C_interp = rbf([[T_input, p_input]])[0]
    else:
        f_interp = interp1d(T_aray, C_aray, kind='linear', fill_value='extrapolate')
        if verbose and (T_input < np.min(T_aray) or T_input > np.max(T_aray)):
            print(f"⚠️ Экстраполяция по температуре T={T_input} для '{name_ohl}'")
        C_interp = f_interp(T_input)

    return float(C_interp)
def find_temp_st_g(T_og, T_st_g, T_ohl, alpha_ohl, kpd_r, q_kon, q_l,delta_st, lambda_st_vn, d, number, ind_smena, lambda_st_vn_1,verbose=False):
    """
    Расчёт температуры горячей и холодной стенки камеры охлаждения ЖРД, с учётом смены материала в определённой точке.

    Используется баланс тепловых потоков (конвекция, излучение) и сопротивление теплопроводности.

    Args:
        T_og (float): Температура газа.
        T_st_g (list): Начальная температура горячей стенки (по точкам).
        T_ohl (list): Температура охладителя.
        alpha_ohl (list): Коэффициент теплоотдачи охладителя.
        kpd_r (list): КПД теплообмена.
        q_kon (list): Конвективный тепловой поток [МВт/м²].
        q_l (list): Лучистый тепловой поток [МВт/м²].
        delta_st (float): Толщина стенки (м).
        lambda_st_vn (list): Теплопроводность для материала до точки смены.
        d (float): Диаметр канала.
        number (list): Индексы расчётных точек.
        ind_smena (int): Индекс, начиная с которого материал сменяется.
        lambda_st_vn_1 (list): Теплопроводность для нового материала после `ind_smena`.
        verbose (bool): Печатать значения для отладки.

    Returns:
        tuple:
            - T_st_g_itog (list): Расчётная температура горячей стенки.
            - T_st_ohl (list): Температура охладителя у стенки.
            - lambda_mat (list): Теплопроводность в каждой точке.
    """
    T_st_g_itog = []
    T_st_ohl = []
    lambda_mat = []
    k = 0

    # Температурные зависимости теплопроводности (перенести в constants при необходимости)
    x_m = np.array([273, 373, 473, 573, 723, 798, 873, 923, 1023, 1073])
    y_m = np.array([220, 240, 260, 280, 300, 310, 315, 310, 300, 292])

    x_s = np.array([373, 423, 473, 523, 573, 623, 673, 723, 773, 823, 873, 923, 973, 1023, 1073, 1123, 1173])
    y_s = np.array([16.5, 16.8, 17.4, 18.1, 19, 19.9, 20.5, 21.6, 22.1, 23, 23.6, 24.1, 25.1, 25.6, 26.6, 27.8, 28.8])

    for t_ohl_i, alpha_i, eta_i, q_k_i, q_l_i, i in zip(
        T_ohl, alpha_ohl, kpd_r, q_kon, q_l, number
    ):
        q_k_i *= 1e6
        q_l_i *= 1e6
        q_total = q_k_i + q_l_i

        lambda_list = lambda_st_vn if i <= ind_smena else lambda_st_vn_1
        lambda_i = lambda_list[i]

        R_cond = delta_st / lambda_i
        R_conv = 1 / (alpha_i * eta_i)
        R_total = R_cond + R_conv

        T_ohl_wall = t_ohl_i
        T_st_g_i = d * (
            T_og / (T_og - T_st_g[i]) +
            T_ohl_wall / (R_total * q_total) +
            q_l_i / q_total
        ) / (
            1 / (T_og - T_st_g[i]) + 1 / (R_total * q_total)
        )

        T_st_g_itog.append(T_st_g_i)

        T_lambda = T_st_g_i
        if i <= ind_smena:
            lambda_val = np.interp(T_lambda, x_m, y_m, left=y_m[0], right=y_m[-1])
        else:
            lambda_val = np.interp(T_lambda, x_s, y_s, left=y_s[0], right=y_s[-1])
        lambda_mat.append(lambda_val)

        T_st_ohl.append(T_st_g_i - (delta_st / lambda_val) * q_total)

        if verbose:
            print(f"[{i}] T_st_g: {T_st_g_i:.2f} K, λ: {lambda_val:.2f}, T_st_ohl: {T_st_ohl[-1]:.2f} K")

        k += 1

    # Экспорт в Excel (опционально)
    df = pd.DataFrame({'T_st_g_itog': T_st_g_itog, 'T_st_ohl': T_st_ohl})
    df.fillna('-').to_excel('table5.xlsx', index=False)

    return T_st_g_itog, T_st_ohl, lambda_mat, lambda_mat
def find_params_ohl(name_ohl):
    """
    Загружает теплофизические свойства охладителя по его названию из Excel-файла.

    Разделение:
    - Сложные охладители (зависят от давления): возвращаются 7 столбцов, включая `p`.
    - Простые охладители: 6 столбцов, без давления.

    Args:
        name_ohl (str): Название охладителя (например, "АТ", "Водород", "НДМГ", и т.д.)

    Returns:
        tuple:
            - Если компонент сложный: (T, p, ρ, C_p, μ, λ, K)
            - Если компонент простой: (T, ρ, C_p, μ, λ, K)
    """
    file_map = {
        "Водород": ("data/Properties of components/Hydrogen_Features.xlsx", 7),
        "АТ": ("data/Properties of components/Nitrogen-tetraoxide-Features.xlsx", 7),
        "Кислород": ("data/Properties of components/Oxygen-Features.xlsx", 7),
        "Гелий": ("data/Properties of components/Helium-Features.xlsx", 7),
        "Аммиак": ("data/Properties of components/Ammonia-Features.xlsx", 7),
        "Метан": ("data/Properties of components/Methane-Features.xlsx", 7),
        "НДМГ": ("data/Properties of components/Hydrazine_Features.xlsx", 6),
        "Аэрозин-50": ("data/Properties of components/Aerosin-50-Features.xlsx", 6),
        "Вода": ("data/Properties of components/Water_Features.xlsx", 6),
        "Керосин-Т1": ("data/Properties of components/Kerosene-T-1-Features.xlsx", 6),
        "Этанол": ("data/Properties of components/Ethanol-Features.xlsx", 6),
    }

    if name_ohl not in file_map:
        raise ValueError(f"Неизвестный охладитель: {name_ohl}")

    file_path, num_columns = file_map[name_ohl]

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл не найден: {file_path}")

    workbook = load_workbook(file_path)
    sheet = workbook.active

    columns = [[] for _ in range(num_columns)]
    start_row = 2  # пропускаем заголовок

    for col in range(1, num_columns + 1):
        row = start_row
        while True:
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is None:
                break
            columns[col - 1].append(cell_value)
            row += 1

    # Распаковка
    if num_columns == 7:
        T, p, rho, Cp, mu, lamb, K = columns
        return T, p, rho, Cp, mu, lamb, K
    else:
        T, rho, Cp, mu, lamb, K = columns
        return T, rho, Cp, mu, lamb, K
def find_nearest_index(array, value):
    """
    Находит индекс ближайшего по значению элемента в массиве к заданному числу.

    Args:
        array (list of float): Список чисел, среди которых ведётся поиск.
        value (float): Значение, к которому ищется ближайшее.

    Returns:
        int: Индекс ближайшего элемента.

    Raises:
        ValueError: Если передан пустой массив.
    """
    if not array:
        raise ValueError("Массив пустой — нельзя найти ближайший элемент.")
    return min(range(len(array)), key=lambda i: abs(array[i] - value))
def reflect_tail(x, y, n):
    """
    Зеркально отражает последние `n` точек массива `x, y` и добавляет отражённые значения в конец.

    Отражение выполняется относительно последней точки хвоста по оси X,
    создавая симметричное продолжение графика.

    Args:
        x (list of float): Исходный массив координат по X.
        y (list of float): Исходный массив координат по Y.
        n (int): Количество точек для отражения (с конца).

    Returns:
        tuple:
            - x_new (list of float): Расширенный массив X с отражённой частью.
            - y_new (list of float): Расширенный массив Y с отражённой частью.

    Raises:
        ValueError: Если n больше длины входных массивов.
    """
    if n > len(x) or n > len(y):
        raise ValueError("Количество точек для отражения превышает длину массива.")

    x_tail = x[-n:]
    y_tail = y[-n:]

    x_mirror = x_tail[-1]
    x_reflected = [2 * x_mirror - xi for xi in reversed(x_tail)]
    y_reflected = list(reversed(y_tail))

    x_new = x + x_reflected
    y_new = y + y_reflected

    return x_new, y_new
def clone_temp_st(array_1, array_2, n, X=None):
    """
    Создаёт сглаженный симметричный хвост в массивах температуры (горячей и холодной стенки),
    усредняя зеркальные точки относительно центра последнего участка.

    Алгоритм:
    - Делит последние `n` точек на две половины.
    - Усредняет попарно симметричные точки.
    - Склеивает с основной частью массива.

    Args:
        array_1 (list of float): Первый температурный профиль (например, T_st_g).
        array_2 (list of float): Второй профиль (например, T_st_ohl).
        n (int): Количество точек в хвосте, которые подлежат зеркальному усреднению (n должно быть чётным).
        X (list, optional): Не используется (можно удалить или добавить использование).

    Returns:
        tuple:
            - array_1_smoothed (list of float): Обновлённый профиль с симметричным хвостом.
            - array_2_smoothed (list of float): Аналогично, для второго массива.
    """
    def smooth_symmetric_tail(arr, n):
        start = arr[:len(arr) - n]
        tail = arr[-n:]
        mid = n // 2
        left = tail[:mid]
        right = tail[mid:][::-1]
        smoothed = [(a + b) * 0.5 for a, b in zip(left, right)]
        symmetric_tail = smoothed + smoothed[::-1]
        return start + symmetric_tail

    if n % 2 != 0:
        raise ValueError("n должно быть чётным для корректного зеркального усреднения.")
    if len(array_1) < n or len(array_2) < n:
        raise ValueError("Длина массивов должна быть больше n.")

    array_1_smoothed = smooth_symmetric_tail(array_1, n)
    array_2_smoothed = smooth_symmetric_tail(array_2, n)

    return array_1_smoothed, array_2_smoothed